from __future__ import annotations

from dataclasses import dataclass, field

from django.contrib.auth import get_user_model

from .models import Category, Shayari


LANGUAGE_ALIASES = {
    'hi': Shayari.LANG_HI,
    'hindi': Shayari.LANG_HI,
    'en': Shayari.LANG_EN,
    'english': Shayari.LANG_EN,
    'ur': Shayari.LANG_UR,
    'urdu': Shayari.LANG_UR,
}

EXPECTED_COLUMNS = {'title', 'text', 'language', 'category', 'author'}
REQUIRED_COLUMNS = {'text'}
DEFAULT_TITLE = 'Untitled'
DEFAULT_CATEGORY = 'General'
DEFAULT_LANGUAGE = Shayari.LANG_HI
HEADER_ALIASES = {
    'title': 'title',
    'shayarititle': 'title',
    'text': 'text',
    'shayari': 'text',
    'shayaritext': 'text',
    'content': 'text',
    'body': 'text',
    'language': 'language',
    'lang': 'language',
    'category': 'category',
    'cat': 'category',
    'author': 'author',
    'authorname': 'author',
    'username': 'author',
    'user': 'author',
}


@dataclass
class ImportResult:
    created: int = 0
    skipped: int = 0
    warnings: list[str] = field(default_factory=list)


def _clean(value):
    if value is None:
        return ''
    return str(value).strip()


def _header_key(value):
    raw = _clean(value).lower()
    return ''.join(ch for ch in raw if ch.isalnum())


def _resolve_language(raw_language: str):
    return LANGUAGE_ALIASES.get(raw_language.strip().lower())


def _resolve_author(raw_author: str, default_author, result: ImportResult, row_num: int):
    raw_author = raw_author.strip()
    if not raw_author:
        return default_author

    User = get_user_model()
    user = User.objects.filter(username__iexact=raw_author).first()
    if user:
        return user

    user = User.objects.filter(email__iexact=raw_author).first()
    if user:
        return user

    result.warnings.append(
        f"Row {row_num}: author '{raw_author}' not found, used '{default_author.username}'."
    )
    return default_author


def import_shayari_xlsx(file_obj, default_author, approve=False):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ValueError('openpyxl is required to import .xlsx files.') from exc

    workbook = load_workbook(filename=file_obj, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header_row = None
    header_row_num = 1
    for row_number in range(1, 6):
        candidate = next(rows, None)
        if not candidate:
            continue
        normalized = {
            HEADER_ALIASES.get(_header_key(value))
            for value in candidate
            if _header_key(value)
        }
        if normalized & REQUIRED_COLUMNS:
            header_row = candidate
            header_row_num = row_number
            break

    if not header_row:
        raise ValueError(
            'Could not find header row. Expected columns: Title, Text, Language, Category, Author.'
        )

    column_indices = {}
    for idx, header in enumerate(header_row):
        key = HEADER_ALIASES.get(_header_key(header))
        if key and key in EXPECTED_COLUMNS and key not in column_indices:
            column_indices[key] = idx

    missing = REQUIRED_COLUMNS - set(column_indices.keys())
    if missing:
        missing_str = ', '.join(sorted(missing))
        raise ValueError(f'Missing required columns: {missing_str}.')

    result = ImportResult()

    for row_num, row in enumerate(rows, start=header_row_num + 1):
        if row is None:
            continue

        title = (
            _clean(row[column_indices['title']])
            if 'title' in column_indices and column_indices['title'] < len(row)
            else ''
        )
        text = (
            _clean(row[column_indices['text']])
            if 'text' in column_indices and column_indices['text'] < len(row)
            else ''
        )
        language_raw = (
            _clean(row[column_indices['language']])
            if 'language' in column_indices and column_indices['language'] < len(row)
            else ''
        )
        category_name = (
            _clean(row[column_indices['category']])
            if 'category' in column_indices and column_indices['category'] < len(row)
            else ''
        )
        author_raw = (
            _clean(row[column_indices['author']])
            if 'author' in column_indices and column_indices['author'] < len(row)
            else ''
        )

        if not any([title, text, language_raw, category_name, author_raw]):
            continue

        if not text:
            result.skipped += 1
            result.warnings.append(f'Row {row_num}: missing text.')
            continue

        title = title or DEFAULT_TITLE
        if len(title) > 200:
            result.skipped += 1
            result.warnings.append(f'Row {row_num}: title exceeds 200 characters.')
            continue

        language = _resolve_language(language_raw) if language_raw else DEFAULT_LANGUAGE
        if not language:
            result.skipped += 1
            result.warnings.append(
                f"Row {row_num}: invalid language '{language_raw}' (use Hindi/English/Urdu)."
            )
            continue

        category_name = category_name or DEFAULT_CATEGORY

        category = Category.objects.filter(name__iexact=category_name).first()
        if not category:
            category = Category.objects.create(name=category_name)
        author = _resolve_author(author_raw, default_author, result, row_num)

        Shayari.objects.create(
            title=title,
            text=text,
            language=language,
            category=category,
            author=author,
            approved=approve,
        )
        result.created += 1

    workbook.close()
    return result
